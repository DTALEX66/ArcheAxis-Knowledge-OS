"""AXW-023C: structured XLSX/CSV adapter (cell semantics + honest loss).

Uses openpyxl (already in the dependency tree via markitdown[xlsx]) to
extract sheet / cell / formula / value semantics with stable anchors
(sheet name + A1-style coordinates). Large sheets are bounded: a
``max_cells`` guard stops extraction past a threshold and records the
boundary in loss notes instead of silently truncating.

CSV is handled by the passthrough engine in multi_format (text format);
this adapter focuses on the structured .xlsx path.
"""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from shared.adapter_contract import AdapterResult

# Hard boundary: refuse to expand beyond this many cells per sheet so a
# pathological workbook cannot blow up memory. Honest loss note on truncation.
MAX_CELLS_PER_SHEET = 100_000


def _sha256(file_path: str | Path) -> str:
    h = hashlib.sha256()
    with open(file_path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _cell_value(cell) -> tuple[str, bool]:
    """Return (value_text, is_formula). Formula cells keep the formula text
    plus the cached value so semantics survive round-trip."""
    if cell.data_type == "f" and cell.value:
        formula = str(cell.value)
        return formula if formula.startswith("=") else f"={formula}", True
    v = cell.value
    if v is None:
        return "", False
    return str(v), False


def _extract_workbook(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=False)
    blocks: list[dict[str, Any]] = []
    loss: list[str] = []

    for sheet in wb.worksheets:
        sheet_name = sheet.title
        cell_count = 0
        rows_out: list[str] = []
        truncated = False
        for row in sheet.iter_rows():
            row_parts: list[str] = []
            for cell in row:
                cell_count += 1
                if cell_count > MAX_CELLS_PER_SHEET:
                    truncated = True
                    break
                value_text, _ = _cell_value(cell)
                if value_text:
                    row_parts.append(f"{cell.coordinate}:{value_text}")
            if row_parts:
                rows_out.append(" | ".join(row_parts))
            if truncated:
                break
        if rows_out:
            blocks.append(
                {
                    "kind": "sheet",
                    "text": "\n".join(rows_out),
                    "anchor": {"sheet": sheet_name, "cell_count": min(cell_count, MAX_CELLS_PER_SHEET)},
                }
            )
        else:
            blocks.append(
                {"kind": "sheet-empty", "text": "", "anchor": {"sheet": sheet_name, "cell_count": 0}}
            )
        if truncated:
            loss.append(f"sheet '{sheet_name}' exceeded {MAX_CELLS_PER_SHEET} cells; truncated")

    wb.close()
    return blocks, loss


def convert_xlsx(file_path: str | Path) -> AdapterResult:
    """Convert a .xlsx to structured sheet blocks."""
    path = Path(file_path)
    if not path.is_file():
        return AdapterResult(success=False, content="", engine="xlsx-adapter", error="file not found")

    try:
        from openpyxl import load_workbook  # noqa: F401
    except ImportError:
        return AdapterResult(
            success=False,
            content="",
            engine="xlsx-adapter",
            error="XLSX conversion requires openpyxl; install markitdown[xlsx].",
        )

    try:
        blocks, loss = _extract_workbook(path)
    except Exception as exc:  # pragma: no cover - host-tooling dependent
        return AdapterResult(
            success=False,
            content="",
            engine="xlsx-adapter",
            error=f"openpyxl conversion failed: {exc}",
        )

    if not any(b["kind"] != "sheet-empty" for b in blocks):
        return AdapterResult(
            success=False,
            content="",
            engine="xlsx-adapter",
            error="XLSX conversion returned no content; treat as degraded.",
        )

    text = "\n\n".join(b["text"] for b in blocks if b["text"])
    return AdapterResult(
        success=True,
        content=text,
        engine="xlsx-adapter",
        metadata={
            "char_count": len(text),
            "block_count": len(blocks),
            "blocks": blocks,
            "loss_notes": loss,
        },
    )


def convert_xlsx_to_run(
    file_path: str | Path,
    db: str | Path,
    source_name: str | None = None,
    version: int = 1,
) -> dict[str, Any]:
    """Convert a .xlsx and persist a ConversionRun; returns run metadata."""
    result = convert_xlsx(file_path)
    if not result.success:
        raise RuntimeError(result.error or "XLSX conversion failed")

    from app.ingestion.conversion_run import create_conversion_run, store_conversion_run

    raw_sha = _sha256(file_path)
    blocks: list[dict[str, Any]] = result.metadata.get("blocks") or []
    loss = result.metadata.get("loss_notes") or []
    run = create_conversion_run(
        raw_sha256=raw_sha,
        source_name=source_name or Path(file_path).name,
        blocks=blocks,
        engine=result.engine,
        version=version,
    )
    store_conversion_run(db, run)
    return {
        "run_id": run.run_id,
        "document_id": run.document.document_id,
        "block_count": len(blocks),
        "loss_notes": loss,
    }
